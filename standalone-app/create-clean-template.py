"""
テンプレートファイルから共有数式を削除してクリーンなテンプレートを作成し、Base64エンコードするスクリプト

使い方:
1. Pythonをインストール (3.7以上)
2. pip install openpyxl
3. python create-clean-template.py
"""

import openpyxl
from openpyxl.utils import get_column_letter
import base64
import os

def create_clean_template():
    print('テンプレートファイルを読み込み中...')

    template_path = os.path.join(os.path.dirname(__file__), 'tyouzai_excel_v2.xlsx')
    workbook = openpyxl.load_workbook(template_path)
    worksheet = workbook.active

    print('共有数式をクリア中...')

    # すべての行の数式をクリア（6行目から最大1000行まで）
    for row_num in range(6, 1001):
        for col_num in range(1, 21):  # A列からT列まで（20列）
            cell = worksheet.cell(row=row_num, column=col_num)
            # セルに数式があれば削除
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                cell.value = None

    print('クリーンなテンプレートを保存中...')
    clean_template_path = os.path.join(os.path.dirname(__file__), 'tyouzai_excel_v2_clean.xlsx')
    workbook.save(clean_template_path)

    print('Base64エンコード中...')
    # ファイルを読み込んでBase64エンコード
    with open(clean_template_path, 'rb') as f:
        file_data = f.read()
        base64_data = base64.b64encode(file_data).decode('utf-8')

    # Base64文字列をファイルに保存
    base64_path = os.path.join(os.path.dirname(__file__), 'template_base64.txt')
    with open(base64_path, 'w', encoding='utf-8') as f:
        f.write(base64_data)

    # JavaScriptファイルとして保存
    js_path = os.path.join(os.path.dirname(__file__), 'template-data.js')
    with open(js_path, 'w', encoding='utf-8') as f:
        f.write('// クリーンなテンプレートファイル (Base64エンコード済み)\n')
        f.write('const TEMPLATE_BASE64 = \'')
        f.write(base64_data)
        f.write('\';\n')

    print('\n完了しました！')
    print(f'クリーンなテンプレート: {clean_template_path}')
    print(f'Base64ファイル: {base64_path}')
    print(f'JavaScriptファイル: {js_path}')
    print(f'Base64サイズ: {len(base64_data)} 文字')
    print(f'元のファイルサイズ: {len(file_data)} バイト')

if __name__ == '__main__':
    try:
        create_clean_template()
    except Exception as e:
        print(f'❌ エラーが発生しました: {e}')
        import traceback
        traceback.print_exc()
