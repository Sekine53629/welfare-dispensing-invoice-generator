#!/usr/bin/env python3
"""
Excelテンプレート作成スクリプト（テーブル構造なし）
ヘッダー情報のみを含む最小限のテンプレートを作成
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import base64

# 新しいワークブック作成
wb = Workbook()
ws = wb.active
ws.title = "Sheet1"

# ヘッダー情報（1-9行目）
ws['A1'] = '調剤券請求書'
ws['A1'].font = Font(size=16, bold=True)
ws.merge_cells('A1:M1')
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

ws['A3'] = '請求年月:'
ws['A4'] = '薬局名:'
ws['A5'] = '医療機関コード:'

ws['B3'] = '2025年2月分'
ws['B4'] = ''  # 動的に設定
ws['B5'] = ''  # 動的に設定

# 9行目まで空白
for row in range(6, 10):
    ws.row_dimensions[row].height = 15

# 10行目: テーブルヘッダー（ExcelJSで上書きされる）
# ※ここではヘッダーを作成せず、ExcelJSに任せる
# 空行として残す
ws.row_dimensions[10].height = 20

print("✅ クリーンテンプレート作成完了")

# ファイル保存
template_path = 'standalone-app/template-clean-no-table.xlsx'
wb.save(template_path)
print(f"💾 保存先: {template_path}")

# Base64エンコード
with open(template_path, 'rb') as f:
    template_bytes = f.read()
    template_base64 = base64.b64encode(template_bytes).decode('utf-8')

# template-data.jsに出力
js_content = f"""/**
 * Excelテンプレートデータ（Base64エンコード）
 * Version: 2.3.3 - テーブル構造なしバージョン
 */

const TEMPLATE_BASE64 = '{template_base64}';

// ブラウザ環境で使用
if (typeof window !== 'undefined') {{
    window.TEMPLATE_BASE64 = TEMPLATE_BASE64;
}}

// Node.js環境で使用
if (typeof module !== 'undefined' && module.exports) {{
    module.exports = TEMPLATE_BASE64;
}}
"""

with open('standalone-app/template-data.js', 'w', encoding='utf-8') as f:
    f.write(js_content)

print(f"✅ template-data.js 更新完了")
print(f"📊 Base64サイズ: {len(template_base64)} 文字")
